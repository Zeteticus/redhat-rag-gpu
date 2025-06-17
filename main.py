#!/usr/bin/env python3
"""
Red Hat Documentation RAG Backend - Multi-Format Support
FastAPI-based backend for intelligent document search and retrieval
Optimized for RHEL with GPU Acceleration and Multi-Format Document Processing

Supported Formats:
- PDF (PyMuPDF for better extraction)
- DOCX (python-docx)
- TXT (native text files)
- XLSX/XLS (openpyxl, xlrd)
- EPUB (ebooklib)
- MOBI (python-kindle)
- HTML (BeautifulSoup)
- RTF (striprtf)
- ODT (odfpy)
- CSV (built-in csv module)
"""

import os
import re
import json
import time
import logging
import hashlib
import shutil
import torch
import mimetypes
import zipfile
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Any, Union, Tuple
from dataclasses import dataclass, asdict
from contextlib import asynccontextmanager
from fastapi import Request

import uvicorn
import numpy as np
from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# Multi-format document processing libraries
try:
    import fitz  # PyMuPDF - better PDF processing
    PDF_AVAILABLE = True
except ImportError:
    import PyPDF2  # Fallback to PyPDF2
    PDF_AVAILABLE = False
    
try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    import xlrd
    XLS_AVAILABLE = True
except ImportError:
    XLS_AVAILABLE = False

try:
    import ebooklib
    from ebooklib import epub
    EPUB_AVAILABLE = True
except ImportError:
    EPUB_AVAILABLE = False

try:
    from bs4 import BeautifulSoup
    HTML_AVAILABLE = True
except ImportError:
    HTML_AVAILABLE = False

try:
    from striprtf import rtf_to_text
    RTF_AVAILABLE = True
except ImportError:
    RTF_AVAILABLE = False

try:
    from odf import text, teletype
    from odf.opendocument import load
    ODT_AVAILABLE = True
except ImportError:
    ODT_AVAILABLE = False

try:
    import kindle_unpack
    MOBI_AVAILABLE = True
except ImportError:
    MOBI_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuration
class Config:
    DOCUMENTS_DIR = os.getenv('DOCUMENTS_DIR', './documents')
    CHROMA_DB_PATH = os.getenv('CHROMA_DB_PATH', './vectordb')
    MODEL_NAME = os.getenv('EMBEDDING_MODEL', 'all-MiniLM-L6-v2')
    CHUNK_SIZE = int(os.getenv('CHUNK_SIZE', '500'))
    CHUNK_OVERLAP = int(os.getenv('CHUNK_OVERLAP', '50'))
    MAX_RESULTS = int(os.getenv('MAX_RESULTS', '20'))
    MIN_CONFIDENCE = float(os.getenv('MIN_CONFIDENCE', '0.3'))
    GPU_BATCH_SIZE = int(os.getenv('GPU_BATCH_SIZE', '32'))
    MAX_FILE_SIZE = int(os.getenv('MAX_FILE_SIZE', '100')) * 1024 * 1024  # 100MB default

    # Supported file extensions
    SUPPORTED_EXTENSIONS = {
        '.pdf': 'PDF Document',
        '.docx': 'Microsoft Word Document',
        '.doc': 'Microsoft Word Document (Legacy)',
        '.txt': 'Plain Text',
        '.xlsx': 'Microsoft Excel Spreadsheet',
        '.xls': 'Microsoft Excel Spreadsheet (Legacy)',
        '.epub': 'EPUB eBook',
        '.mobi': 'MOBI eBook',
        '.html': 'HTML Document',
        '.htm': 'HTML Document',
        '.rtf': 'Rich Text Format',
        '.odt': 'OpenDocument Text',
        '.csv': 'Comma-Separated Values'
    }

    # Red Hat specific patterns
    RHEL_VERSION_PATTERN = r'(?:RHEL|Red Hat Enterprise Linux)\s*(\d+(?:\.\d+)?)'
    CATEGORY_KEYWORDS = {
        'installation': ['install', 'setup', 'deployment', 'bootstrap'],
        'networking': ['network', 'ip', 'dns', 'dhcp', 'firewall', 'iptables'],
        'security': ['security', 'selinux', 'authentication', 'ssl', 'tls', 'encryption'],
        'storage': ['storage', 'filesystem', 'disk', 'lvm', 'raid', 'mount'],
        'virtualization': ['kvm', 'qemu', 'libvirt', 'virtual', 'hypervisor'],
        'containers': ['container', 'podman', 'docker', 'kubernetes', 'openshift'],
        'troubleshooting': ['troubleshoot', 'debug', 'error', 'problem', 'issue']
    }

# Pydantic Models
class SearchRequest(BaseModel):
    query: str = Field(..., min_length=1, max_length=500)
    filters: Optional[Dict[str, Any]] = Field(default_factory=dict)
    max_results: Optional[int] = Field(default=10, ge=1, le=50)
    min_confidence: Optional[float] = Field(default=0.3, ge=0.0, le=1.0)

class SearchResult(BaseModel):
    id: str
    title: str
    content: str
    source: str
    confidence: float
    category: str
    version: str
    tags: List[str]
    page: int
    section: str
    file_type: str
    metadata: Dict[str, Any]

class SearchResponse(BaseModel):
    results: List[SearchResult]
    total_count: int
    response_time_ms: int
    query: str
    filters_applied: Dict[str, Any]

class StatsResponse(BaseModel):
    total_documents: int
    total_chunks: int
    total_queries: int
    avg_response_time_ms: float
    system_status: str
    last_updated: str
    categories: Dict[str, int]
    versions: Dict[str, int]
    file_types: Dict[str, int]
    gpu_enabled: bool
    gpu_info: Optional[str] = None
    supported_formats: Dict[str, bool]

@dataclass
class DocumentChunk:
    """Represents a chunk of text from a document"""
    id: str
    content: str
    source: str
    page: int
    section: str
    title: str
    category: str
    version: str
    tags: List[str]
    file_type: str
    metadata: Dict[str, Any]
    embedding: Optional[np.ndarray] = None

class MultiFormatExtractor:
    """Handles text extraction from multiple file formats"""
    
    def __init__(self):
        self.supported_formats = {}
        self._register_extractors()
    
    def _register_extractors(self):
        """Register available extractors based on installed libraries"""
        if PDF_AVAILABLE:
            self.supported_formats['.pdf'] = self._extract_pdf_pymupdf
        else:
            self.supported_formats['.pdf'] = self._extract_pdf_pypdf2
            
        if DOCX_AVAILABLE:
            self.supported_formats['.docx'] = self._extract_docx
            
        if XLSX_AVAILABLE:
            self.supported_formats['.xlsx'] = self._extract_xlsx
            
        if XLS_AVAILABLE:
            self.supported_formats['.xls'] = self._extract_xls
            
        if EPUB_AVAILABLE:
            self.supported_formats['.epub'] = self._extract_epub
            
        if HTML_AVAILABLE:
            self.supported_formats['.html'] = self._extract_html
            self.supported_formats['.htm'] = self._extract_html
            
        if RTF_AVAILABLE:
            self.supported_formats['.rtf'] = self._extract_rtf
            
        if ODT_AVAILABLE:
            self.supported_formats['.odt'] = self._extract_odt
            
        if MOBI_AVAILABLE:
            self.supported_formats['.mobi'] = self._extract_mobi
        
        # Always available formats
        self.supported_formats['.txt'] = self._extract_txt
        self.supported_formats['.csv'] = self._extract_csv
        
        logger.info(f"Registered extractors for: {list(self.supported_formats.keys())}")

    def get_file_type(self, file_path: Path) -> str:
        """Detect file type from extension and MIME type"""
        extension = file_path.suffix.lower()
        if extension in self.supported_formats:
            return extension
        
        # Try MIME type detection
        mime_type, _ = mimetypes.guess_type(str(file_path))
        if mime_type:
            mime_to_ext = {
                'application/pdf': '.pdf',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
                'application/vnd.ms-excel': '.xls',
                'text/plain': '.txt',
                'text/html': '.html',
                'text/csv': '.csv',
                'application/epub+zip': '.epub',
                'application/rtf': '.rtf'
            }
            return mime_to_ext.get(mime_type, extension)
        
        return extension

    def can_process(self, file_path: Path) -> bool:
        """Check if file can be processed"""
        file_type = self.get_file_type(file_path)
        return file_type in self.supported_formats

    def extract_text(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from file based on its type"""
        file_type = self.get_file_type(file_path)
        
        if file_type not in self.supported_formats:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        try:
            extractor = self.supported_formats[file_type]
            return extractor(file_path)
        except Exception as e:
            logger.error(f"Error extracting from {file_path}: {str(e)}")
            # Try generic text extraction as fallback
            return self._extract_text_fallback(file_path)

    def _extract_pdf_pymupdf(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from PDF using PyMuPDF (better than PyPDF2)"""
        try:
            doc = fitz.open(str(file_path))
            pages = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                
                if text.strip():
                    # Extract additional metadata
                    blocks = page.get_text("dict")
                    images = page.get_images()
                    
                    pages.append({
                        'page': page_num + 1,
                        'text': text,
                        'source': file_path.name,
                        'metadata': {
                            'image_count': len(images),
                            'block_count': len(blocks.get('blocks', [])),
                            'char_count': len(text)
                        }
                    })
            
            doc.close()
            logger.info(f"Extracted {len(pages)} pages from PDF: {file_path.name}")
            return pages
            
        except Exception as e:
            logger.error(f"PyMuPDF extraction failed for {file_path}: {str(e)}")
            return self._extract_pdf_pypdf2(file_path)

    def _extract_pdf_pypdf2(self, file_path: Path) -> List[Dict[str, Any]]:
        """Fallback PDF extraction using PyPDF2"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                pages = []
                
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    if text.strip():
                        pages.append({
                            'page': page_num,
                            'text': text,
                            'source': file_path.name,
                            'metadata': {'char_count': len(text)}
                        })
                
                logger.info(f"Extracted {len(pages)} pages from PDF (PyPDF2): {file_path.name}")
                return pages
                
        except Exception as e:
            logger.error(f"PyPDF2 extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_docx(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from DOCX files"""
        try:
            doc = DocxDocument(str(file_path))
            paragraphs = []
            page_num = 1
            current_section = "Document"
            
            full_text = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    full_text.append(paragraph.text)
            
            # Group paragraphs into logical pages (every ~500 words)
            words_per_page = 500
            current_words = 0
            current_page_text = []
            
            for para_text in full_text:
                word_count = len(para_text.split())
                
                if current_words + word_count > words_per_page and current_page_text:
                    # Save current page
                    paragraphs.append({
                        'page': page_num,
                        'text': '\n'.join(current_page_text),
                        'source': file_path.name,
                        'section': current_section,
                        'metadata': {
                            'word_count': current_words,
                            'paragraph_count': len(current_page_text)
                        }
                    })
                    
                    # Start new page
                    page_num += 1
                    current_page_text = [para_text]
                    current_words = word_count
                else:
                    current_page_text.append(para_text)
                    current_words += word_count
            
            # Add remaining text
            if current_page_text:
                paragraphs.append({
                    'page': page_num,
                    'text': '\n'.join(current_page_text),
                    'source': file_path.name,
                    'section': current_section,
                    'metadata': {
                        'word_count': current_words,
                        'paragraph_count': len(current_page_text)
                    }
                })
            
            logger.info(f"Extracted {len(paragraphs)} logical pages from DOCX: {file_path.name}")
            return paragraphs
            
        except Exception as e:
            logger.error(f"DOCX extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_xlsx(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from XLSX files"""
        try:
            workbook = load_workbook(str(file_path), read_only=True)
            sheets = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        rows_data.append(' | '.join(row_text))
                
                if rows_data:
                    sheets.append({
                        'page': len(sheets) + 1,
                        'text': '\n'.join(rows_data),
                        'source': file_path.name,
                        'section': f"Sheet: {sheet_name}",
                        'metadata': {
                            'sheet_name': sheet_name,
                            'row_count': len(rows_data),
                            'max_column': sheet.max_column
                        }
                    })
            
            workbook.close()
            logger.info(f"Extracted {len(sheets)} sheets from XLSX: {file_path.name}")
            return sheets
            
        except Exception as e:
            logger.error(f"XLSX extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_xls(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from XLS files"""
        try:
            workbook = xlrd.open_workbook(str(file_path))
            sheets = []
            
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                rows_data = []
                
                for row_idx in range(sheet.nrows):
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.value:
                            row_values.append(str(cell.value))
                    if row_values:
                        rows_data.append(' | '.join(row_values))
                
                if rows_data:
                    sheets.append({
                        'page': sheet_idx + 1,
                        'text': '\n'.join(rows_data),
                        'source': file_path.name,
                        'section': f"Sheet: {sheet.name}",
                        'metadata': {
                            'sheet_name': sheet.name,
                            'row_count': sheet.nrows,
                            'col_count': sheet.ncols
                        }
                    })
            
            logger.info(f"Extracted {len(sheets)} sheets from XLS: {file_path.name}")
            return sheets
            
        except Exception as e:
            logger.error(f"XLS extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_epub(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from EPUB files"""
        try:
            book = epub.read_epub(str(file_path))
            chapters = []
            chapter_num = 1
            
            for item in book.get_items():
                if item.get_type() == ebooklib.ITEM_DOCUMENT:
                    content = item.get_content().decode('utf-8')
                    
                    # Parse HTML content
                    if HTML_AVAILABLE:
                        soup = BeautifulSoup(content, 'html.parser')
                        text = soup.get_text()
                    else:
                        # Basic HTML tag removal
                        import re
                        text = re.sub(r'<[^>]+>', '', content)
                    
                    if text.strip():
                        chapters.append({
                            'page': chapter_num,
                            'text': text,
                            'source': file_path.name,
                            'section': f"Chapter {chapter_num}",
                            'metadata': {
                                'item_id': item.get_id(),
                                'item_name': item.get_name(),
                                'char_count': len(text)
                            }
                        })
                        chapter_num += 1
            
            logger.info(f"Extracted {len(chapters)} chapters from EPUB: {file_path.name}")
            return chapters
            
        except Exception as e:
            logger.error(f"EPUB extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_html(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from HTML files"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
            
            if HTML_AVAILABLE:
                soup = BeautifulSoup(content, 'html.parser')
                
                # Remove script and style elements
                for script in soup(["script", "style"]):
                    script.extract()
                
                text = soup.get_text()
                title = soup.title.string if soup.title else file_path.stem
            else:
                # Basic HTML tag removal
                import re
                text = re.sub(r'<[^>]+>', '', content)
                title = file_path.stem
            
            # Clean up whitespace
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            if text.strip():
                return [{
                    'page': 1,
                    'text': text,
                    'source': file_path.name,
                    'section': title,
                    'metadata': {
                        'title': title,
                        'char_count': len(text),
                        'file_size': file_path.stat().st_size
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"HTML extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_rtf(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from RTF files"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                rtf_content = file.read()
            
            text = rtf_to_text(rtf_content)
            
            if text.strip():
                return [{
                    'page': 1,
                    'text': text,
                    'source': file_path.name,
                    'section': "RTF Document",
                    'metadata': {
                        'char_count': len(text),
                        'file_size': file_path.stat().st_size
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"RTF extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_odt(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from ODT files"""
        try:
            doc = load(str(file_path))
            paragraphs = doc.getElementsByType(text.P)
            
            text_content = []
            for paragraph in paragraphs:
                para_text = teletype.extractText(paragraph)
                if para_text.strip():
                    text_content.append(para_text)
            
            full_text = '\n'.join(text_content)
            
            if full_text.strip():
                return [{
                    'page': 1,
                    'text': full_text,
                    'source': file_path.name,
                    'section': "ODT Document",
                    'metadata': {
                        'paragraph_count': len(text_content),
                        'char_count': len(full_text)
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"ODT extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_mobi(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from MOBI files"""
        try:
            # This is a placeholder - MOBI extraction is complex
            # Would need kindle_unpack or similar library
            logger.warning(f"MOBI extraction not fully implemented for {file_path}")
            return []
            
        except Exception as e:
            logger.error(f"MOBI extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_txt(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from plain text files"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
            
            if content.strip():
                return [{
                    'page': 1,
                    'text': content,
                    'source': file_path.name,
                    'section': "Text Document",
                    'metadata': {
                        'char_count': len(content),
                        'line_count': len(content.splitlines()),
                        'file_size': file_path.stat().st_size
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"Text extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_csv(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract text from CSV files"""
        try:
            rows_data = []
            
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                # Try to detect delimiter
                sample = file.read(1024)
                file.seek(0)
                
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.reader(file, delimiter=delimiter)
                
                for row_num, row in enumerate(reader, 1):
                    if row and any(cell.strip() for cell in row):
                        row_text = ' | '.join(str(cell).strip() for cell in row if str(cell).strip())
                        if row_text:
                            rows_data.append(f"Row {row_num}: {row_text}")
            
            if rows_data:
                return [{
                    'page': 1,
                    'text': '\n'.join(rows_data),
                    'source': file_path.name,
                    'section': "CSV Data",
                    'metadata': {
                        'row_count': len(rows_data),
                        'delimiter': delimiter,
                        'file_size': file_path.stat().st_size
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"CSV extraction failed for {file_path}: {str(e)}")
            return []

    def _extract_text_fallback(self, file_path: Path) -> List[Dict[str, Any]]:
        """Fallback text extraction for unknown formats"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read(10000)  # Read first 10KB
            
            # Basic cleanup
            content = re.sub(r'[^\w\s\.\,\!\?\;\:\-\(\)]', ' ', content)
            content = re.sub(r'\s+', ' ', content).strip()
            
            if content and len(content) > 50:  # Must have substantial content
                return [{
                    'page': 1,
                    'text': content,
                    'source': file_path.name,
                    'section': "Unknown Format (Fallback)",
                    'metadata': {
                        'extraction_method': 'fallback',
                        'char_count': len(content),
                        'file_type': file_path.suffix
                    }
                }]
            
            return []
            
        except Exception as e:
            logger.error(f"Fallback extraction failed for {file_path}: {str(e)}")
            return []

class GPUVectorStore:
    """
    A vector storage implementation with GPU acceleration.
    Uses basic file storage and numpy for vector operations.
    """

    def __init__(self, db_path: str = './vectordb'):
        self.db_path = Path(db_path)
        self.vectors_file = self.db_path / 'vectors.npy'
        self.metadata_file = self.db_path / 'metadata.json'
        self.documents_file = self.db_path / 'documents.json'
        self.ids_file = self.db_path / 'ids.json'

        self.model = None
        self.device = "cpu"
        self.gpu_info = None
        self._stats = {
            'total_queries': 0,
            'response_times': [],
            'last_updated': datetime.now()
        }

        # Create directory if it doesn't exist
        os.makedirs(self.db_path, exist_ok=True)

    def ensure_storage_directory(self):
        """Ensure the storage directory exists and is writable"""
        try:
            # Create directory if it doesn't exist
            os.makedirs(self.db_path, exist_ok=True)

            # Check if directory is writable by attempting to create a test file
            test_file = self.db_path / "write_test.txt"
            with open(test_file, 'w') as f:
                f.write("Test write access")
            os.remove(test_file)

            # Create subdirectories for better organization
            os.makedirs(self.db_path / "data", exist_ok=True)
            os.makedirs(self.db_path / "metadata", exist_ok=True)

            return True
        except Exception as e:
            logger.error(f"Storage directory error: {str(e)}")
            return False

    async def initialize(self):
        """Initialize the vector store with GPU support if available"""
        try:
            # Check storage directory
            if not self.ensure_storage_directory():
                logger.error("Failed to ensure storage directory is accessible")
                raise Exception("Storage directory is not accessible for read/write")

            # Check for GPU availability
            if torch.cuda.is_available():
                self.device = f"cuda:{torch.cuda.current_device()}"
                self.gpu_info = f"{torch.cuda.get_device_name(0)} (CUDA {torch.version.cuda})"
                logger.info(f"GPU detected: {self.gpu_info}")
                logger.info(f"Using device: {self.device}")
            else:
                self.device = "cpu"
                logger.warning("No GPU detected, falling back to CPU")

            # Initialize the model with the appropriate device
            logger.info(f"Loading embedding model: {Config.MODEL_NAME} on {self.device}")
            self.model = SentenceTransformer(Config.MODEL_NAME, device=self.device)

            # Initialize empty data structures if files don't exist
            if not self.vectors_file.exists():
                self._save_vectors(np.array([]).reshape(0, 384))  # Default embedding size for all-MiniLM-L6-v2

            if not self.metadata_file.exists():
                self._save_metadata([])

            if not self.documents_file.exists():
                self._save_documents([])

            if not self.ids_file.exists():
                self._save_ids([])

            # Load existing data
            if self.vectors_file.exists() and self.metadata_file.exists() and self.documents_file.exists() and self.ids_file.exists():
                logger.info("Loading existing vector store data...")

                # Verify data integrity
                try:
                    vectors = self._load_vectors()
                    metadata = self._load_metadata()
                    documents = self._load_documents()
                    ids = self._load_ids()

                    if len(metadata) != len(documents) or len(metadata) != len(ids):
                        logger.warning("Data integrity issue detected - lengths don't match")
                        if vectors.size > 0 and vectors.shape[0] != len(metadata):
                            logger.warning("Vector count doesn't match metadata count")

                    logger.info(f"Loaded {len(metadata)} items from persistent storage")
                except Exception as e:
                    logger.error(f"Error loading data: {str(e)}")
                    logger.warning("Creating new empty data store")

                    # Initialize empty data structures
                    self._save_vectors(np.array([]).reshape(0, 384))
                    self._save_metadata([])
                    self._save_documents([])
                    self._save_ids([])

            logger.info(f"Vector store initialized successfully on {self.device}")

        except Exception as e:
            logger.error(f"Failed to initialize vector store: {str(e)}")
            raise

    def _load_vectors(self) -> np.ndarray:
        """Load vectors from file"""
        if self.vectors_file.exists():
            return np.load(str(self.vectors_file), allow_pickle=True)
        return np.array([]).reshape(0, 384)

    def _save_vectors(self, vectors: np.ndarray):
        """Save vectors to file with fix for numpy.save() .npy extension behavior"""
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(self.vectors_file), exist_ok=True)
            
            # Create temp filename without .npy since numpy will add it
            base_name = str(self.vectors_file).replace('.npy', '')
            temp_base = base_name + '.tmp'  # e.g., /app/vectordb/vectors.tmp
            
            # numpy.save will create vectors.tmp.npy
            np.save(temp_base, vectors)
            
            # The actual temp file created by numpy
            actual_temp_file = temp_base + '.npy'  # e.g., /app/vectordb/vectors.tmp.npy
            
            # Move the actual temp file to final location
            shutil.move(actual_temp_file, str(self.vectors_file))
            
            logger.info(f"Saved {vectors.shape[0]} vectors to {self.vectors_file}")
            
        except Exception as e:
            # Cleanup any temp files that might exist
            cleanup_files = [temp_base + '.npy', temp_base] if 'temp_base' in locals() else []
            for cleanup_file in cleanup_files:
                if os.path.exists(cleanup_file):
                    try:
                        os.remove(cleanup_file)
                    except:
                        pass
            
            # Fallback: Try direct save without temp file
            try:
                logger.warning(f"Temp file method failed ({e}), trying direct save...")
                # Remove .npy from target since numpy will add it back
                direct_target = str(self.vectors_file).replace('.npy', '')
                np.save(direct_target, vectors)
                logger.info("Direct save successful")
            except Exception as e2:
                logger.error(f"Direct save also failed: {e2}")
                raise e2

    def _load_metadata(self) -> List[Dict[str, Any]]:
        """Load metadata from file"""
        if self.metadata_file.exists():
            with open(self.metadata_file, 'r') as f:
                return json.load(f)
        return []

    def _save_metadata(self, metadata: List[Dict[str, Any]]):
        """Save metadata to file with error handling and atomic writes"""
        try:
            # Save to temporary file first
            temp_file = str(self.metadata_file) + ".tmp"
            with open(temp_file, 'w') as f:
                json.dump(metadata, f)

            # Then rename to actual file
            shutil.move(temp_file, str(self.metadata_file))

            logger.info(f"Saved metadata for {len(metadata)} items")
        except Exception as e:
            logger.error(f"Error saving metadata: {str(e)}")
            raise

    def _load_documents(self) -> List[str]:
        """Load documents from file"""
        if self.documents_file.exists():
            with open(self.documents_file, 'r') as f:
                return json.load(f)
        return []

    def _save_documents(self, documents: List[str]):
        """Save documents to file with error handling and atomic writes"""
        try:
            # Save to temporary file first
            temp_file = str(self.documents_file) + ".tmp"
            with open(temp_file, 'w') as f:
                json.dump(documents, f)

            # Then rename to actual file
            shutil.move(temp_file, str(self.documents_file))

            logger.info(f"Saved {len(documents)} documents")
        except Exception as e:
            logger.error(f"Error saving documents: {str(e)}")
            raise

    def _load_ids(self) -> List[str]:
        """Load IDs from file"""
        if self.ids_file.exists():
            with open(self.ids_file, 'r') as f:
                return json.load(f)
        return []

    def _save_ids(self, ids: List[str]):
        """Save IDs to file with error handling and atomic writes"""
        try:
            # Save to temporary file first
            temp_file = str(self.ids_file) + ".tmp"
            with open(temp_file, 'w') as f:
                json.dump(ids, f)

            # Then rename to actual file
            shutil.move(temp_file, str(self.ids_file))

            logger.info(f"Saved {len(ids)} IDs")
        except Exception as e:
            logger.error(f"Error saving IDs: {str(e)}")
            raise

    def add_chunks(self, chunks):
        """Add document chunks to the vector store with GPU acceleration"""
        if not chunks:
            return

        try:
            # Load existing data
            vectors = self._load_vectors()
            metadata = self._load_metadata()
            documents = self._load_documents()
            ids = self._load_ids()

            # Generate embeddings with GPU acceleration if available
            texts = [chunk.content for chunk in chunks]

            # Get current batch size based on available memory
            batch_size = Config.GPU_BATCH_SIZE  # Default batch size
            if torch.cuda.is_available():
                # For large GPUs, increase batch size
                gpu_mem = torch.cuda.get_device_properties(0).total_memory / (1024**3)  # GB
                if gpu_mem > 16:
                    batch_size = 128
                elif gpu_mem > 8:
                    batch_size = 64
                logger.info(f"Using GPU with batch size {batch_size}")

            # Generate embeddings in batches
            all_embeddings = []
            for i in range(0, len(texts), batch_size):
                batch_texts = texts[i:i+batch_size]
                logger.info(f"Processing batch {i//batch_size + 1}/{(len(texts) + batch_size - 1)//batch_size}")
                batch_embeddings = self.model.encode(
                    batch_texts,
                    show_progress_bar=True,
                    batch_size=batch_size
                )
                all_embeddings.append(batch_embeddings)
                # Free up CUDA memory
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()

            # Combine batches
            if len(all_embeddings) > 1:
                new_embeddings = np.vstack(all_embeddings)
            else:
                new_embeddings = all_embeddings[0]

            # Prepare data
            new_ids = [chunk.id for chunk in chunks]
            new_metadata = []
            new_documents = []

            for chunk in chunks:
                new_metadata.append({
                    'source': chunk.source,
                    'page': chunk.page,
                    'section': chunk.section,
                    'title': chunk.title,
                    'category': chunk.category,
                    'version': chunk.version,
                    'file_type': chunk.file_type,
                    'tags': json.dumps(chunk.tags),
                    **chunk.metadata
                })
                new_documents.append(chunk.content)

            # Append to existing data
            if vectors.size > 0:
                vectors = np.vstack([vectors, new_embeddings])
            else:
                vectors = new_embeddings

            metadata.extend(new_metadata)
            documents.extend(new_documents)
            ids.extend(new_ids)

            # Save updated data
            self._save_vectors(vectors)
            self._save_metadata(metadata)
            self._save_documents(documents)
            self._save_ids(ids)

            # Update stats
            self._stats['last_updated'] = datetime.now()

            logger.info(f"Added {len(chunks)} chunks to vector store using {self.device}")

        except Exception as e:
            logger.error(f"Error adding chunks to vector store: {str(e)}")
            raise

    def search(self, query: str, filters: Dict[str, Any] = None,
               max_results: int = 10, min_confidence: float = 0.3) -> List[Dict[str, Any]]:
        """Search for relevant documents using GPU acceleration for query embedding"""
        start_time = time.time()

        try:
            # Load data
            vectors = self._load_vectors()
            metadata = self._load_metadata()
            documents = self._load_documents()
            ids = self._load_ids()

            if vectors.size == 0:
                logger.warning("Vector store is empty")
                return []

            # Generate query embedding on GPU if available
            query_embedding = self.model.encode([query])[0]

            # Calculate cosine similarity (optimized for larger datasets)
            # Move operation to GPU if possible, otherwise use CPU
            if self.device != "cpu" and vectors.shape[0] > 10000:
                # For very large datasets, use GPU for similarity calculation
                logger.info(f"Using GPU for similarity calculation with {vectors.shape[0]} vectors")

                # Convert to torch tensors on GPU
                vectors_tensor = torch.tensor(vectors, device=self.device)
                query_tensor = torch.tensor(query_embedding, device=self.device)

                # Normalize tensors
                vectors_tensor = vectors_tensor / torch.norm(vectors_tensor, dim=1, keepdim=True)
                query_tensor = query_tensor / torch.norm(query_tensor)

                # Calculate similarities
                similarities = torch.matmul(vectors_tensor, query_tensor).cpu().numpy()
            else:
                # Use numpy for smaller datasets
                dot_products = np.dot(vectors, query_embedding)
                magnitudes = np.linalg.norm(vectors, axis=1) * np.linalg.norm(query_embedding)
                similarities = dot_products / magnitudes

            # Apply filters
            filtered_indices = []
            for i, meta in enumerate(metadata):
                if filters:
                    if filters.get('category') and meta.get('category') != filters.get('category'):
                        continue
                    if filters.get('version') and meta.get('version') != filters.get('version'):
                        continue
                    if filters.get('file_type') and meta.get('file_type') != filters.get('file_type'):
                        continue

                confidence = similarities[i]
                if confidence >= min_confidence:
                    filtered_indices.append(i)

            # Sort by similarity and take top results
            sorted_indices = [idx for idx in sorted(filtered_indices, key=lambda i: similarities[i], reverse=True)]

            results = []
            for i in sorted_indices[:max_results]:
                meta = metadata[i]
                results.append({
                    'id': ids[i],
                    'title': meta['title'],
                    'content': documents[i],
                    'source': meta['source'],
                    'confidence': round(float(similarities[i]), 3),
                    'category': meta['category'],
                    'version': meta['version'],
                    'file_type': meta.get('file_type', 'unknown'),
                    'tags': json.loads(meta.get('tags', '[]')),
                    'page': meta['page'],
                    'section': meta['section'],
                    'metadata': {
                        k: v for k, v in meta.items()
                        if k not in ['title', 'source', 'category', 'version', 'tags', 'page', 'section', 'file_type']
                    }
                })

            # Update stats
            response_time = int((time.time() - start_time) * 1000)
            self._stats['total_queries'] += 1
            self._stats['response_times'].append(response_time)
            if len(self._stats['response_times']) > 100:
                self._stats['response_times'] = self._stats['response_times'][-100:]

            logger.info(f"Search completed: {len(results)} results in {response_time}ms using {self.device}")
            return results

        except Exception as e:
            logger.error(f"Search error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")

    def get_stats(self) -> Dict[str, Any]:
        """Get vector store statistics"""
        try:
            # Load metadata
            metadata = self._load_metadata()

            # Calculate stats
            categories = {}
            versions = {}
            file_types = {}

            for meta in metadata:
                cat = meta.get('category', 'unknown')
                ver = meta.get('version', 'unknown')
                ft = meta.get('file_type', 'unknown')
                
                categories[cat] = categories.get(cat, 0) + 1
                versions[ver] = versions.get(ver, 0) + 1
                file_types[ft] = file_types.get(ft, 0) + 1

            avg_response_time = (
                sum(self._stats['response_times']) / len(self._stats['response_times'])
                if self._stats['response_times'] else 0
            )

            return {
                'total_chunks': len(metadata),
                'total_queries': self._stats['total_queries'],
                'avg_response_time_ms': round(avg_response_time, 2),
                'categories': categories,
                'versions': versions,
                'file_types': file_types,
                'last_updated': self._stats['last_updated'].isoformat(),
                'gpu_enabled': self.device != "cpu",
                'gpu_info': self.gpu_info
            }

        except Exception as e:
            logger.error(f"Error getting stats: {str(e)}")
            return {}

class DocumentProcessor:
    """Handles multi-format document processing and text extraction"""

    def __init__(self):
        self.config = Config()
        self.extractor = MultiFormatExtractor()

    def detect_rhel_version(self, text: str) -> str:
        """Detect RHEL version from document text"""
        match = re.search(self.config.RHEL_VERSION_PATTERN, text, re.IGNORECASE)
        if match:
            version = match.group(1)
            major_version = version.split('.')[0]
            return f"rhel{major_version}"

        # Additional detection for RHEL 10 with specific keywords
        if re.search(r'RHEL\s*10|Red Hat Enterprise Linux\s*10', text, re.IGNORECASE) or \
           (re.search(r'RHEL|Red Hat Enterprise Linux', text, re.IGNORECASE) and
            re.search(r'(2025|latest release|newest version)', text, re.IGNORECASE)):
            return "rhel10"

        return "unknown"

    def categorize_content(self, text: str) -> str:
        """Categorize content based on keywords"""
        text_lower = text.lower()
        scores = {}

        for category, keywords in self.config.CATEGORY_KEYWORDS.items():
            score = sum(1 for keyword in keywords if keyword in text_lower)
            if score > 0:
                scores[category] = score

        return max(scores, key=scores.get) if scores else "general"

    def extract_tags(self, text: str) -> List[str]:
        """Extract relevant tags from text"""
        tags = set()
        text_lower = text.lower()

        # Technical terms
        tech_patterns = [
            r'\b(systemctl|systemd|firewalld|selinux|podman|docker)\b',
            r'\b(yum|dnf|rpm|subscription-manager)\b',
            r'\b(ssh|http|https|ftp|nfs|samba)\b',
            r'\b(tcp|udp|ip|dns|dhcp)\b'
        ]

        for pattern in tech_patterns:
            matches = re.findall(pattern, text_lower)
            tags.update(matches)

        return list(tags)[:10]  # Limit to 10 tags

    def extract_section_title(self, text: str) -> str:
        """Extract section title from text"""
        lines = text.split('\n')
        for line in lines[:5]:  # Check first 5 lines
            line = line.strip()
            if line and (line.isupper() or line.startswith('Chapter') or
                        line.startswith('Section') or len(line.split()) <= 8):
                return line[:100]  # Limit length
        return "Content"

    def chunk_text(self, text: str, chunk_size: int = None, overlap: int = None) -> List[str]:
        """Split text into overlapping chunks"""
        chunk_size = chunk_size or self.config.CHUNK_SIZE
        overlap = overlap or self.config.CHUNK_OVERLAP

        words = text.split()
        chunks = []

        for i in range(0, len(words), chunk_size - overlap):
            chunk = ' '.join(words[i:i + chunk_size])
            if chunk.strip():
                chunks.append(chunk)

        return chunks

    def process_document(self, file_path: Path) -> List[DocumentChunk]:
        """Process a document of any supported format into chunks"""
        
        if not self.extractor.can_process(file_path):
            logger.warning(f"Unsupported file format: {file_path}")
            return []

        try:
            # Extract text using appropriate extractor
            pages = self.extractor.extract_text(file_path)
            if not pages:
                logger.warning(f"No content extracted from {file_path}")
                return []

            chunks = []
            doc_title = file_path.stem.replace('-', ' ').replace('_', ' ').title()
            file_type = self.extractor.get_file_type(file_path)

            # Detect document-level metadata
            full_text = ' '.join([page['text'] for page in pages])
            doc_version = self.detect_rhel_version(full_text)
            doc_category = self.categorize_content(full_text)

            for page_data in pages:
                page_text = page_data['text']
                page_chunks = self.chunk_text(page_text)

                for i, chunk_text in enumerate(page_chunks):
                    chunk_id = hashlib.md5(
                        f"{file_path.name}_{page_data['page']}_{i}".encode()
                    ).hexdigest()

                    chunk = DocumentChunk(
                        id=chunk_id,
                        content=chunk_text,
                        source=file_path.name,
                        page=page_data['page'],
                        section=page_data.get('section', self.extract_section_title(chunk_text)),
                        title=doc_title,
                        category=self.categorize_content(chunk_text) or doc_category,
                        version=self.detect_rhel_version(chunk_text) or doc_version,
                        file_type=file_type,
                        tags=self.extract_tags(chunk_text),
                        metadata={
                            'file_size': file_path.stat().st_size,
                            'processed_at': datetime.now().isoformat(),
                            'chunk_index': i,
                            'total_chunks': len(page_chunks),
                            'extraction_metadata': page_data.get('metadata', {}),
                            **page_data.get('metadata', {})
                        }
                    )
                    chunks.append(chunk)

            logger.info(f"Processed {file_path.name} ({file_type}): {len(chunks)} chunks created")
            return chunks

        except Exception as e:
            logger.error(f"Error processing {file_path}: {str(e)}")
            return []

class DocumentManager:
    """Manages document processing and indexing for multiple formats"""

    def __init__(self, vector_store: GPUVectorStore):
        self.config = Config()
        self.processor = DocumentProcessor()
        self.vector_store = vector_store
        self.processed_files = set()

        # Ensure documents directory exists
        os.makedirs(self.config.DOCUMENTS_DIR, exist_ok=True)

    async def scan_and_process_documents(self):
        """Scan documents directory and process all supported file types"""
        docs_path = Path(self.config.DOCUMENTS_DIR)
        
        # Find all supported files
        supported_files = []
        for ext in self.config.SUPPORTED_EXTENSIONS.keys():
            pattern = f"*{ext}"
            files = list(docs_path.glob(pattern))
            supported_files.extend(files)

        logger.info(f"Found {len(supported_files)} supported files")

        for file_path in supported_files:
            if file_path.name not in self.processed_files:
                await self.process_document(file_path)
                self.processed_files.add(file_path.name)

    async def process_document(self, file_path: Path):
        """Process a single document"""
        try:
            logger.info(f"Processing document: {file_path.name}")
            chunks = self.processor.process_document(file_path)

            if chunks:
                self.vector_store.add_chunks(chunks)
                logger.info(f"Successfully processed {file_path.name}")
            else:
                logger.warning(f"No content extracted from {file_path.name}")

        except Exception as e:
            logger.error(f"Error processing {file_path.name}: {str(e)}")

    async def add_document(self, file_content: bytes, filename: str):
        """Add a new document from uploaded content"""
        try:
            # Check file size
            if len(file_content) > self.config.MAX_FILE_SIZE:
                raise HTTPException(status_code=413, detail=f"File size exceeds maximum limit of {self.config.MAX_FILE_SIZE // (1024*1024)}MB")

            # Check if file type is supported
            file_ext = Path(filename).suffix.lower()
            if file_ext not in self.config.SUPPORTED_EXTENSIONS:
                raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_ext}. Supported types: {list(self.config.SUPPORTED_EXTENSIONS.keys())}")

            # First try to write to a temporary location
            temp_dir = "/tmp/uploads"
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = Path(temp_dir) / filename

            with open(temp_path, 'wb') as f:
                f.write(file_content)

            # Then try to copy to the actual documents directory
            target_path = Path(self.config.DOCUMENTS_DIR) / filename
            try:
                shutil.copy2(temp_path, target_path)
                file_path = target_path
                logger.info(f"Document saved to permanent location: {target_path}")
            except Exception as e:
                # If copy fails, use the temp file
                logger.warning(f"Could not copy to documents dir: {str(e)}, using temp file")
                file_path = temp_path

            # Process the document
            await self.process_document(file_path)
            self.processed_files.add(filename)

            return {
                "message": f"Document {filename} processed successfully",
                "file_type": self.config.SUPPORTED_EXTENSIONS.get(file_ext, "Unknown"),
                "size": len(file_content)
            }

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error adding document {filename}: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Failed to process document: {str(e)}")

    def get_supported_formats(self) -> Dict[str, bool]:
        """Get information about supported file formats"""
        formats = {}
        
        # Check which formats are actually available
        extractor = MultiFormatExtractor()
        for ext, description in self.config.SUPPORTED_EXTENSIONS.items():
            formats[ext] = {
                "description": description,
                "available": ext in extractor.supported_formats
            }
        
        return formats

# Global instances
vector_store = None
doc_manager = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan management"""
    # Startup
    logger.info("Starting Red Hat Documentation RAG Backend with Multi-Format Support and GPU Acceleration")

    global vector_store
    vector_store = GPUVectorStore(db_path=Config.CHROMA_DB_PATH)
    await vector_store.initialize()

    global doc_manager
    doc_manager = DocumentManager(vector_store)

    # Process existing documents
    await doc_manager.scan_and_process_documents()

    logger.info("Backend initialization complete")
    yield

    # Shutdown
    logger.info("Shutting down backend")
    # Free GPU memory if applicable
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

# FastAPI application
app = FastAPI(
    title="Red Hat Documentation RAG API (Multi-Format, GPU-Accelerated)",
    description="Intelligent search and retrieval for Red Hat system administration documentation with multi-format support and GPU acceleration",
    version="2.0.0",
    lifespan=lifespan
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static files (frontend)
if os.path.exists("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def serve_frontend():
    """Serve the frontend HTML"""
    static_path = Path("static/index.html")
    if static_path.exists():
        return FileResponse("static/index.html")
    else:
        return JSONResponse(
            content={
                "message": "Red Hat Documentation RAG API (Multi-Format, GPU-Accelerated)", 
                "docs": "/docs",
                "supported_formats": list(Config.SUPPORTED_EXTENSIONS.keys())
            },
            status_code=200
        )

@app.post("/api/search", response_model=SearchResponse)
async def search_documents(request: SearchRequest):
    """Search through documentation"""
    start_time = time.time()

    try:
        results = vector_store.search(
            query=request.query,
            filters=request.filters,
            max_results=request.max_results,
            min_confidence=request.min_confidence
        )

        # Convert raw results to SearchResult model
        search_results = []
        for result in results:
            search_results.append(SearchResult(**result))

        response_time = int((time.time() - start_time) * 1000)

        return SearchResponse(
            results=search_results,
            total_count=len(search_results),
            response_time_ms=response_time,
            query=request.query,
            filters_applied=request.filters
        )

    except Exception as e:
        logger.error(f"Search API error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/stats", response_model=StatsResponse)
async def get_stats():
    """Get system statistics"""
    try:
        stats = vector_store.get_stats()

        # Count total documents by type
        docs_path = Path(Config.DOCUMENTS_DIR)
        total_docs = 0
        file_type_counts = {}
        
        for ext in Config.SUPPORTED_EXTENSIONS.keys():
            files = list(docs_path.glob(f"*{ext}"))
            count = len(files)
            total_docs += count
            if count > 0:
                file_type_counts[ext] = count

        return StatsResponse(
            total_documents=total_docs,
            total_chunks=stats.get('total_chunks', 0),
            total_queries=stats.get('total_queries', 0),
            avg_response_time_ms=stats.get('avg_response_time_ms', 0),
            system_status="healthy",
            last_updated=stats.get('last_updated', datetime.now().isoformat()),
            categories=stats.get('categories', {}),
            versions=stats.get('versions', {}),
            file_types=stats.get('file_types', {}),
            gpu_enabled=stats.get('gpu_enabled', False),
            gpu_info=stats.get('gpu_info'),
            supported_formats=doc_manager.get_supported_formats() if doc_manager else {}
        )

    except Exception as e:
        logger.error(f"Stats API error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/documents/upload")
async def upload_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Upload a new document (supports multiple formats)"""
    
    # Check file extension
    file_ext = Path(file.filename).suffix.lower() if file.filename else ""
    if file_ext not in Config.SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Unsupported file type: {file_ext}. Supported formats: {list(Config.SUPPORTED_EXTENSIONS.keys())}"
        )

    try:
        content = await file.read()
        result = await doc_manager.add_document(content, file.filename)

        return JSONResponse(
            content=result,
            status_code=201
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/{filename}")
@app.head("/api/documents/{filename}")
async def get_document(filename: str, request: Request):
    """Retrieve a document file for inline viewing in browser"""
    try:
        # Security: prevent directory traversal attacks
        if ".." in filename or "/" in filename or "\\" in filename:
            logger.warning(f"Invalid filename attempted: {filename}")
            raise HTTPException(status_code=400, detail="Invalid filename")

        # Log the request for debugging
        logger.info(f"Document {request.method} request for: {filename}")

        # Check if file type is supported for serving
        file_ext = Path(filename).suffix.lower()
        
        # First try the main documents directory
        file_path = Path(Config.DOCUMENTS_DIR) / filename
        logger.debug(f"Checking main documents path: {file_path}")

        if not file_path.exists():
            # If not found, try the temporary uploads directory
            temp_path = Path("/tmp/uploads") / filename
            logger.debug(f"Checking temp path: {temp_path}")
            if temp_path.exists():
                file_path = temp_path
                logger.debug(f"Found in temp directory: {temp_path}")
            else:
                logger.error(f"Document not found: {filename}")

                # List available files for debugging
                docs_path = Path(Config.DOCUMENTS_DIR)
                if docs_path.exists():
                    available_files = []
                    for ext in Config.SUPPORTED_EXTENSIONS.keys():
                        available_files.extend(list(docs_path.glob(f"*{ext}")))
                    logger.debug(f"Available files: {[f.name for f in available_files]}")

                raise HTTPException(status_code=404, detail=f"Document '{filename}' not found")

        # Verify it's a supported file type
        if file_ext not in Config.SUPPORTED_EXTENSIONS:
            logger.warning(f"Unsupported file type requested: {filename}")
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_ext}")

        # Check file permissions
        if not os.access(file_path, os.R_OK):
            logger.error(f"Cannot read file: {file_path}")
            raise HTTPException(status_code=403, detail="File access denied")

        # Determine MIME type
        mime_type, _ = mimetypes.guess_type(filename)
        if not mime_type:
            mime_type = 'application/octet-stream'

        # For HEAD requests, just return success without file content
        if request.method == "HEAD":
            logger.info(f"HEAD request successful for: {filename}")
            from fastapi import Response
            return Response(
                status_code=200,
                headers={
                    "Content-Type": mime_type,
                    "Content-Length": str(file_path.stat().st_size),
                    "Content-Disposition": f"inline; filename={filename}",
                    "Cache-Control": "public, max-age=3600",
                    "X-Content-Type-Options": "nosniff"
                }
            )

        # For GET requests, return the actual file
        logger.info(f"Serving file: {file_path} (size: {file_path.stat().st_size} bytes)")
        return FileResponse(
            path=str(file_path),
            media_type=mime_type,
            filename=filename,
            headers={
                "Content-Disposition": f"inline; filename={filename}",
                "Cache-Control": "public, max-age=3600",
                "X-Content-Type-Options": "nosniff"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error serving document {filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error serving document: {str(e)}")

@app.get("/api/documents")
async def list_documents():
    """List all available documents"""
    try:
        docs_path = Path(Config.DOCUMENTS_DIR)
        all_files = []

        # Check for all supported file types
        for ext in Config.SUPPORTED_EXTENSIONS.keys():
            files = list(docs_path.glob(f"*{ext}"))
            all_files.extend(files)

        # Also check temp directory
        temp_path = Path("/tmp/uploads")
        if temp_path.exists():
            for ext in Config.SUPPORTED_EXTENSIONS.keys():
                temp_files = list(temp_path.glob(f"*{ext}"))
                # Only include temp files that aren't in the main directory
                main_filenames = [f.name for f in all_files]
                all_files.extend([f for f in temp_files if f.name not in main_filenames])

        documents = []
        for file_path in all_files:
            stat = file_path.stat()
            file_ext = file_path.suffix.lower()
            
            documents.append({
                'filename': file_path.name,
                'size': stat.st_size,
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'processed': file_path.name in doc_manager.processed_files,
                'location': 'temporary' if '/tmp/' in str(file_path) else 'permanent',
                'file_type': file_ext,
                'description': Config.SUPPORTED_EXTENSIONS.get(file_ext, 'Unknown')
            })

        return {"documents": documents, "total_count": len(documents)}

    except Exception as e:
        logger.error(f"List documents error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/formats")
async def get_supported_formats():
    """Get information about supported file formats"""
    try:
        formats = doc_manager.get_supported_formats() if doc_manager else {}
        
        return {
            "supported_formats": formats,
            "total_formats": len(formats),
            "available_formats": len([f for f in formats.values() if f.get("available", False)])
        }
    
    except Exception as e:
        logger.error(f"Get formats error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/documents/reprocess")
async def reprocess_documents():
    """Reprocess all documents"""
    try:
        doc_manager.processed_files.clear()
        await doc_manager.scan_and_process_documents()

        return {"message": "Document reprocessing initiated"}

    except Exception as e:
        logger.error(f"Reprocess error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/gpu-info")
async def gpu_info():
    """Get detailed GPU information"""
    gpu_data = {
        "gpu_available": torch.cuda.is_available(),
        "device_count": torch.cuda.device_count() if torch.cuda.is_available() else 0,
        "current_device": torch.cuda.current_device() if torch.cuda.is_available() else None,
    }

    if torch.cuda.is_available():
        gpu_data["device_name"] = torch.cuda.get_device_name(0)
        gpu_data["cuda_version"] = torch.version.cuda

        # Get memory information
        gpu_data["memory"] = {
            "total": torch.cuda.get_device_properties(0).total_memory / (1024**3),  # GB
            "allocated": torch.cuda.memory_allocated(0) / (1024**3),  # GB
            "cached": torch.cuda.memory_reserved(0) / (1024**3),  # GB
        }

    return gpu_data

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    # Check if vector store directory is writable
    vector_store_ok = False
    try:
        if vector_store and vector_store.ensure_storage_directory():
            vector_store_ok = True
    except:
        pass

    # Check if documents directory is writable
    docs_ok = False
    try:
        test_file = Path(Config.DOCUMENTS_DIR) / ".health_check"
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
        docs_ok = True
    except:
        pass

    # Check GPU status
    gpu_ok = torch.cuda.is_available()
    gpu_info = torch.cuda.get_device_name(0) if gpu_ok else "No GPU available"

    # Check extractor availability
    extractor = MultiFormatExtractor()
    format_availability = {ext: ext in extractor.supported_formats for ext in Config.SUPPORTED_EXTENSIONS.keys()}

    return {
        "status": "healthy" if vector_store_ok and docs_ok else "degraded",
        "timestamp": datetime.now().isoformat(),
        "version": "2.0.0",
        "checks": {
            "vector_store": "ok" if vector_store_ok else "failed",
            "documents_dir": "ok" if docs_ok else "failed",
            "gpu": "ok" if gpu_ok else "not available"
        },
        "gpu_info": gpu_info if gpu_ok else None,
        "supported_formats": format_availability,
        "total_supported": sum(1 for available in format_availability.values() if available)
    }

if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8080,
        reload=False,
        log_level="info"
    )
